from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title('OpenPyXL')
root.iconbitmap('excel_tk/palma.ico')
root.geometry("500x800")

# Create workbook instance
wb = Workbook()
# Load existing workbook
wb = load_workbook('Book1.xlsx')
# Create active worksheet
ws = wb.active

# Create variable for Column A
col_a = ws['A']
col_b = ws['B']


def get_a():
    list = ''
    for cell in col_a:
        # print(cell.value)
        list = f'{list + str(cell.value)}\n'
    label_a.config(text=list)


def get_b():
    list = ''
    for cell in col_b:
        # print(cell.value)
        list = f'{list + str(cell.value)}\n'
    label_b.config(text=list)


btn_a = Button(root, text="Get Column A", command=get_a)
btn_a.pack(pady=20)

label_a = Label(root, text="")
label_a.pack(pady=20)

btn_b = Button(root, text="Get Column B", command=get_b)
btn_b.pack(pady=20)

label_b = Label(root, text="")
label_b.pack(pady=20)

# Add Fred to A8 and B8
ws['A8'] = "Fred"
ws['B8'] = "Meat"

# Save new file
wb.save('Book1.xlsx')

root.mainloop()

# if __name__ == '__main__':
